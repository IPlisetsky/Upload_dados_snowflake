import snowflake.connector
from openpyxl import load_workbook
from snowflake.connector import connect, ProgrammingError
from tkinter.filedialog import askopenfilename
import os
import pandas as pd

def snow_flake_api():  
    print("ABRIU")
 
    snowflake_config = {
    'user': '',
    'password': '',
    'account': '',
    'warehouse': '',
    'database': '',
    'schema': '',  
    'role': ''
    }
 
    con = connect(**snowflake_config)
    cursor = con.cursor()
 
    CaminhoArq_ =
    wb = load_workbook(filename=CaminhoArq_, data_only=True)
    sheet = wb.active
    dados = []
    for row in sheet.iter_rows(2, sheet.max_row): 
        dados.append([str(cell.value).replace(',', '') if cell.value is not None else 'NULL' for cell in row]) 

    total_rows = len(dados)
    batch_size = 10000  
    total_inserted = 0

    for i in range(0, total_rows, batch_size):
        batch = dados[i:i + batch_size]
        placeholders = ', '.join(['%s'] * len(batch[0]))
        query = f"INSERT INTO VALUES ({placeholders})"
        cursor.executemany(query, batch)
        total_inserted += len(batch)
        print(f'Inseridas {total_inserted} de {total_rows} linhas')

    cursor.close()
    con.close()
 
    print('Load da FROTA para o datalake realizado com sucesso!')
snow_flake_api()
 
